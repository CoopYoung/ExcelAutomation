import sys 
import pandas as pd
from openpyxl import load_workbook, Workbook
import math
import datetime
from datetime import timedelta
import re
from office365.sharepoint.client_context import ClientContext
from office365.runtime.auth.user_credential import UserCredential
from office365.sharepoint.files.file import File
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import range_boundaries
from openpyxl.styles import Alignment
from io import BytesIO
from openpyxl.styles.borders import Border, Side

#Need access to sharepoint and the correct document_library
def sharepoint():
    sharepoint_site = 'https://navistarinc.sharepoint.com/sites/BuildOperations'
    document_library = 'ace3f3d7-ed91-49d0-808f-16f6d50694bb'

    # Read credentials from file
    file = 'cred.txt'
    with open(file, 'r') as reader:
        uname = reader.readline().strip()
        pwd = reader.readline().strip()

    credentials = UserCredential(uname, pwd)
    ctx = ClientContext(sharepoint_site).with_credentials(credentials)
    
    # Correct the document URL
    file_url = f"{document_library}"
    
    # Load the Excel file from SharePoint
    target_file = ctx.web.get_file_by_server_relative_url(file_url)
    
    # Download the file content
    with open("Build Ops Master Schedule.xlsm", "wb") as local_file:
        target_file.download(local_file).execute_query()

    # Load the content into an openpyxl workbook
    with open("Build Ops Master Schedule.xlsm", "rb") as file_content:
        bo_workbook = load_workbook(file_content)
    
    bo_sheet = bo_workbook['Vehicle Master Build Schedule']
    return bo_sheet
'''
def sharepoint():

    sharepoint_site = 'https://navistarinc.sharepoint.com/:x:/r/sites/BuildOperations/_layouts/15/Doc.aspx?sourcedoc=%7B89CA8623-8556-4938-A0E5-F3D085E92F3D%7D&file=Build%20Ops%20Master%20Schedule.xlsm&action=default&mobileredirect=true&DefaultItemOpen=1'
    document_library = 'ace3f3d7-ed91-49d0-808f-16f6d50694bb'
    #'/sites/BuildOperations/Build Engineering/Master Schedule'
    #/sites/BuildOperations/Build%20Engineers/Master%20Schedule/Build%20Ops%20Master%20Schedule.xlsm?d=w89ca862385564938a0e5f3d085e92f3d&csf=1&web=1&e=9y9gJM
    file = 'cred.txt'
    with open(file, 'r') as reader:
            uname = reader.readline(1)
            pwd = reader.readline(2)
            reader.close()
    credentials = UserCredential(uname, pwd)
    ctx = ClientContext(sharepoint_site).with_credentials(credentials)
    

    # Load the Excel file from SharePoint
    response = ctx.web.get_file_by_server_relative_url(document_library).download()
    response.execute_query()       

    # Load the content into an openpyxl workbook
    excel_content = BytesIO(response.content)
    bo_workbook = load_workbook(excel_content)
    bo_sheet = bo_workbook['Vehicle Master Build Schedule']
    return bo_sheet
'''
#bo_sheet = sharepoint()

source_file = "Copy of Build Ops Master Schedule.xlsm"
target_file = "EES Test Work Load Projections Master -(Cooper).xlsx"


#Build Ops - read only data
bo_workbook = load_workbook(filename=source_file, read_only=True)
bo_sheet = bo_workbook['Vehicle Master Build Schedule'] 
headers = [cell.value for cell in bo_sheet[9]]



# Logic to get column names and indices
def get_column_index(column_name):
    if column_name in headers:
        return headers.index(column_name)
    else:
        raise ValueError(f"Column '{column_name}' not found in headers")

column_names = ['PRGM', 'WRTS #', 'Batch/Build Phase','Truck ID', 'STATUS', 'BUILD SITE', 'Shake Down Duration (working days)', 'BUILD END/EES START      PLANNED', 'BUILD END/EES START        ACTUAL'] 

'''
PRGM : 0
WRTS #: 1
Truck ID: 2
STATUS: 3
BUILD_SITE: 4
Shake down duration: 5
ees start planned: 6
actual: 7
'''
column_indices = [get_column_index(name) for name in column_names]

def gifc(column): #get index from column
    i = 0
    while i < len(column_names):
       # print(i)
        if column == column_names[i]:
            break
        i += 1
    return i



def normalize_vehicle_id(vehicle_id):
    # Use regex to separate the alphabetic part from the numeric part
    if 'LG1' in vehicle_id:
        return vehicle_id.replace("LG1", "")   # Take care of all weird formats 
    if 'TBD' in vehicle_id:
        return 0
    if 'CV' not in vehicle_id and 'DV' not in vehicle_id and 'CERT' not in vehicle_id: 
        return 0
    match = re.match(r"([A-Za-z]+)(\d+)", vehicle_id)
    if match:
        # Extract alphabetic and numeric parts
        alphabetic_part = match.group(1)
        numeric_part = match.group(2).lstrip('0')  # Remove leading zeros from the numeric part
        return f"{alphabetic_part}{numeric_part}"
    else:
        # If the input doesn't match the pattern, return it as is
        return vehicle_id    

def add_projections(p, entire_prog_list, i):

    if len(entire_prog_list) < 1:
        return i
    
    print(f"\n----- ADDING {p} -----\n")
    handoff_idx = column_names.index('BUILD END/EES START      PLANNED')
    vehicle_id_idx = column_names.index('Truck ID')
    start_row = rows_to_clear[i]
    end_row = rows_to_clear[i] + len(entire_prog_list) - 1

    v = 0

    projection_sheet.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=1)
    top_left_Cell = projection_sheet.cell(row=start_row, column=1)
    top_left_Cell.value = p
    top_left_Cell.alignment = Alignment(horizontal='center', vertical='center')
    projection_sheet.cell(row=top_left_Cell.row, column=1).border = medium_border
    for v, prog in enumerate(entire_prog_list):
        handoff_predict = prog[handoff_idx]
        vehicle_id = prog[vehicle_id_idx]

        current_row = start_row + v

        total_vehicle_lst = [cell for cell in projection_sheet['B'][current_row-1:rows_to_clear[-1]]]

        
        for cell in total_vehicle_lst: #Going through projection sheet cells
            if vehicle_id == None: #Not sure why it would be None but add a check
                continue
            if cell.value == None and not isinstance(cell, MergedCell): #Should not be a MergedCell, this is col B
                projection_sheet.cell(row=cell.row, column=2, value=vehicle_id)
                cell.value = vehicle_id
                print(f"Added { p} : {vehicle_id}")
                add_to_sheet(cell.row, handoff_predict)
                projection_sheet.cell(row=current_row, column=2).border = Border(right=Side(style='medium'))
                break
            
        i += 1
    projection_sheet.cell(row=current_row, column=2).border = Border(bottom=Side(style='medium'), right=Side(style='medium'))
    return i #Need to save the idx so next vehicles cascade

def add_to_sheet(r, target_date): #Given the day started, populate the correct spots on excel sheet
   
    ph = [cell.value for cell in projection_sheet[2]] #Dates throughout the year
    column_index = None
    for index, header in enumerate(ph):
        if isinstance(header, datetime.datetime):
            start_of_week = header 
            end_of_week = header + datetime.timedelta(days = 6)
            if start_of_week <= target_date <= end_of_week:
                if target_date.weekday() == 4: #If it is Friday, count the start as following week
                    print("Updating Friday date")
                    column_index = index + 2
                    
                    projection_sheet.cell(row=r, column=column_index, value=1)
                    projection_sheet.cell(row=r, column=column_index+1, value=1)
                    print(f"Added to row {r}")
                else:
                    column_index = index + 1
                    
                    projection_sheet.cell(row=r, column=column_index, value=1)
                    projection_sheet.cell(row=r, column=column_index+1, value=1)
                
                    print(f"Added to row {r}")
                break

def iterate_program(row):
    #This function provides a 1D list of the program information from Build Ops if it is applicable to EES 
    row_data = [cell.value for cell in bo_sheet[row]]
    program_list = []

    if int(row_data[bo_idx_dict['Shake Down Duration (working days)']]) > 0: #Ensure it is a shakedown (Avoids date problems)
        if row_data[bo_idx_dict['BUILD SITE']] == build_site: #Build site either 'N' or 'A'
            if row_data[bo_idx_dict['STATUS']] != 'Complete' and row_data[bo_idx_dict['STATUS']] != 'Canceled': #Not canceled or Complete
                if isinstance(row_data[bo_idx_dict['BUILD END/EES START      PLANNED']], datetime.date): #Must be datetime
                    #print("Made it to date")
                    if isInPast(row_data[bo_idx_dict['BUILD END/EES START      PLANNED']]) == False: #Only July 2024 to Dec                   
                        program_list = [row_data[bo_idx_dict[col_idx]] for col_idx in column_names]
                        return program_list
    else:
        return None
            
                       
def isInPast(t): #If date is before July 1, don't add 
    if t < datetime.datetime(2024, 7, 1, 0, 0) or t.year > 2024:
        return True
    else:
        return False
def format_dates(prg, handoff_predict):
    
    '''
    Format data from iterate_program to :

    -change to datetime 
    
    '''
    print(f"Formatting dates for {prg} ...")
    
    format_data = "%m/%d/%Y"
    predict_date = []
    actual_date = []
    
    
    for t in handoff_predict: #Convert to format -> datetime
        if t == 'N/A':
            continue
        else:
            predict_date.append(datetime.datetime.strptime(t, format_data))


    return predict_date

    ''' This will be for later, for now we will only do handoff_predict

    
    for i, t in enumerate(handoff_actual):
        if t == None:
            return i
        else:
            actual_date.append(datetime.datetime.strptime(t, format_data))
    '''

def read_target(): #This function will grow or shrink as new vehicles come in 

    i = 0
    for ranges in projection_sheet.iter_cols(min_row=1, max_row=124, min_col=1, max_col=2, values_only=True):
        print("")
        #normalize_vehicle_id(ranges[i])
    
    return list(ranges)

def traverse_bo_sheet(): #Goes through whole BO sheet 
    
    cell_value = []
    
    for row in bo_sheet.iter_rows(min_row=10, max_row=bo_sheet.max_row, min_col=2, max_col=2):
        cell_value = row[0].value
        
        if cell_value in seen_values: 
            value_rows[cell_value].append(row[0].row)
                
        else:
            seen_values.add(cell_value)
            if row[0].value != None:
                value_rows[cell_value] = [row[0].row]
                
    return cell_value

def define_indices():
    i = idx = 0
    bo_idx_dict ={}
    while idx < len(column_names):
            while i < len(headers):
                if headers[i] == column_names[idx]:
                    bo_idx_dict[column_names[idx]] = i
                    i = 0 #Reset after finding 
                    break
                    
                else:
                    i += 1
            idx += 1 
    return bo_idx_dict  

def clear_projection_sheet(projection_sheet):
    start_row = 3
    for rows in projection_sheet.iter_rows(min_row=start_row):
        if rows[3].value != None and "1 Truck" in rows[3].value:
            end_row = rows[3].row
            rows_to_clear = [i for i in range(start_row, end_row)]
            break
    
    for row in rows_to_clear:
        for cell in projection_sheet.iter_cols(min_row=row, max_row=rows_to_clear[-1], min_col=1, max_col=len(headers), values_only=False):
            for c in cell:
                clear_cell(c)
    
    print("CLEARED\nTHE\nPROJECTION\nSHEET\n")
    return rows_to_clear
def clear_cell(cell):
    if isinstance(cell, MergedCell): #Merged cells behave differently, this takes into account
        for merged_range in list(projection_sheet.merged_cells.ranges):
            min_col, min_row, max_col, max_row = range_boundaries(str(merged_range))
            if min_col == 1 and max_col == 1:
                projection_sheet.unmerge_cells(str(merged_range))
            ''' #This 'if' is to clear merged_ranges
            if cell.coordinate in merged_range:
                top_left_cell = merged_range.coord.split(":")[0]
                projection_sheet[top_left_cell].value = None
            '''
    else:
        cell.value = None

def main(): #This function is made to make NPG and ATC easier to populate both 
    
    idx = i = 0
    for prog in value_rows:
        entire_prog_list = []
        for row in value_rows[prog]:
            program_list = iterate_program(row) #Row of vehicle and program must be passed
            
            #Now program_list must be added to the projection worksheet, program order must be kept in mind. 
            #It should be split up by program (AKA the outer loop)
            
            if program_list != None:
                entire_prog_list.append(program_list)

        i = add_projections(prog, entire_prog_list, idx) 
        idx = i

    projection_workbook.save(target_file)
    return

if __name__ == '__main__':

    medium_border = Border(left=Side(style='medium'),
                           right=Side(style='medium'),
                           top=Side(style='medium'),
                           bottom=Side(style='medium'))
    seen_values = set() #This set may need to be globalized
    value_rows = {}
    all_progs = traverse_bo_sheet() #This will be added or shrunk depending on workload
    
    bo_idx_dict = define_indices()
    projection_workbook = load_workbook(filename=target_file)
    
    if len(sys.argv) > 1:
        if sys.argv[1] == 'N':
            build_site = sys.argv[1]
            projection_sheet = projection_workbook['2024 NPG Bay Space(Shakedown)']
            rows_to_clear = clear_projection_sheet(projection_sheet)
            main()
            print(" ***** COMPLETED NPG *****")
            projection_sheet = projection_workbook['2024 ATC Bay Space(Shakedown)']
            build_site = 'A'
            rows_to_clear = clear_projection_sheet(projection_sheet)
            main()
            print(" ***** COMPLETED ATC *****")
        elif sys.argv[1] == 'A':
            build_site = sys.argv[1]
            projection_sheet = projection_workbook['2024 ATC Bay Space(Shakedown)']
            rows_to_clear = clear_projection_sheet(projection_sheet)
            main()
            print(" ***** COMPLETED ATC *****")
    else: #default to ATC if no arguments given
        build_site = 'A'
        projection_sheet = projection_workbook['2024 ATC Bay Space(Shakedown)']
        rows_to_clear = clear_projection_sheet(projection_sheet)
        main()
    
    
    #Projections - read and write data
    
    
    
        

    